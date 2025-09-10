"""
Excel export functionality for financial statements with professional formatting
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import Dict, Any, List
import json
import io
import re

class ExcelExporter:
    def __init__(self):
        self.last_token_usage = None
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.currency_format = '#,##0'
    
    def create_statement_dataframe(self, statement_data: Dict, statement_name: str) -> pd.DataFrame:
        """Convert statement data to pandas DataFrame with SEC filing links"""
        if not statement_data.get('periods') or not statement_data.get('data'):
            return pd.DataFrame()
        
        periods = statement_data['periods']
        data = statement_data['data']
        filing_links = statement_data.get('filing_links', {})
        
        # Filter GAAP reconciliation data to only show GAAP and Non-GAAP line items (no adjustments)
        if statement_name == 'GAAP Reconciliation':
            data = self._filter_gaap_reconciliation_data(data)
        
        # Clean and merge line items using ChatGPT
        data = self._clean_and_merge_line_items(data, periods)
        
        # Create DataFrame
        df_data = {'Line Item': list(data.keys())}
        
        for i, period in enumerate(periods):
            period_values = []
            for line_item in data.keys():
                values = data[line_item]
                # Handle case where values might be a dict or other non-list type
                if isinstance(values, dict):
                    # Skip this line item as it's malformed
                    period_values.append('')
                elif isinstance(values, list):
                    if i < len(values) and values[i] is not None:
                        period_values.append(values[i])
                    else:
                        period_values.append('')
                else:
                    # Single value, use it for first period only
                    if i == 0:
                        period_values.append(values)
                    else:
                        period_values.append('')
            df_data[period] = period_values
        
        df = pd.DataFrame(df_data)
        
        # Add SEC filing links row if available
        if filing_links:
            try:
                with st.expander("🔍 Debug: Excel Export Filing Links", expanded=False):
                    st.write(f"Periods: {periods}")
                    st.write(f"Filing Links Dict: {filing_links}")
            except:
                pass
                
            link_row = {'Line Item': 'SEC Filing Link'}
            for period in periods:
                url = filing_links.get(period, '')
                try:
                    with st.expander("🔍 Debug: Period URL Mapping", expanded=False):
                        st.write(f"Period '{period}' -> URL: '{url}'")
                except:
                    pass
                # Store URL for hyperlink formatting later - ensure it's a string
                link_row[period] = str(url) if url else ''
            
            try:
                with st.expander("🔍 Debug: Final Link Row", expanded=False):
                    st.write(f"Final Link Row: {link_row}")
            except:
                pass
                
            # Add the link row as the first row after headers
            link_df = pd.DataFrame([link_row])
            df = pd.concat([link_df, df], ignore_index=True)
            
            try:
                with st.expander("✅ Debug: DataFrame with Filing Links", expanded=False):
                    st.write(f"Added SEC Filing Link row to DataFrame")
                    st.write(f"DataFrame shape after adding links: {df.shape}")
                    st.write(f"First few rows:")
                    st.dataframe(df.head(3))
            except:
                pass
        
        return df
    
    def _filter_gaap_reconciliation_data(self, data: Dict) -> Dict:
        """Filter GAAP reconciliation data to only show GAAP and Non-GAAP line items (exclude adjustments)"""
        filtered_data = {}
        
        # Patterns for GAAP line items
        gaap_patterns = [
            r'(?i)gaap\s+(?:net\s+)?(?:income|loss|earnings|profit)(?:\s+\(loss\))?',
            r'(?i)(?:net\s+)?(?:income|loss|earnings|profit)(?:\s+\(loss\))?.*gaap',
            r'(?i)gaap\s+(?:income|loss)(?:\s+from\s+operations)?(?:\s+\(loss\))?',
            r'(?i)(?:income|loss)(?:\s+from\s+operations)?(?:\s+\(loss\))?.*gaap',
            r'(?i)gaap\s+(?:gross\s+)?profit',
            r'(?i)(?:gross\s+)?profit.*gaap',
            r'(?i)gaap\s+operating\s+(?:income|loss|profit|margin)(?:\s+\(loss\))?',
            r'(?i)operating\s+(?:income|loss|profit|margin)(?:\s+\(loss\))?.*gaap',
            r'(?i)gaap\s+(?:loss|income)\s+from\s+operations',
            r'(?i)(?:loss|income)\s+from\s+operations.*gaap',
            r'(?i)gaap\s+revenue',
            r'(?i)revenue.*gaap',
            r'(?i)gaap\s+(?:loss|earnings|income)\s+per\s+share(?:\s*,\s*(?:basic|diluted))?',
            r'(?i)(?:loss|earnings|income)\s+per\s+share(?:\s*,\s*(?:basic|diluted))?.*gaap',
            r'(?i)gaap\s+(?:eps|earnings\s+per\s+share|net\s+(?:income|loss)\s+per\s+share)(?:\s*,\s*(?:basic|diluted))?',
            r'(?i)(?:eps|earnings\s+per\s+share|net\s+(?:income|loss)\s+per\s+share)(?:\s*,\s*(?:basic|diluted))?.*gaap',
            r'(?i)gaap\s+(?:basic|diluted)\s+(?:eps|earnings\s+per\s+share|(?:income|loss)\s+per\s+share)',
            r'(?i)(?:basic|diluted)\s+(?:eps|earnings\s+per\s+share|(?:income|loss)\s+per\s+share).*gaap',
            r'(?i)gaap\s+(?:total\s+)?operating\s+(?:expenses?|costs?)',
            r'(?i)(?:total\s+)?operating\s+(?:expenses?|costs?).*gaap',
            r'(?i)gaap\s+(?:research\s+and\s+development|r&d|r\s*&\s*d|research\s+&\s+development|research\s+development|product\s+development)',
            r'(?i)(?:research\s+and\s+development|r&d|r\s*&\s*d|research\s+&\s+development|research\s+development|product\s+development).*gaap',
            r'(?i)gaap\s+(?:sales\s+and\s+marketing|marketing)',
            r'(?i)(?:sales\s+and\s+marketing|marketing).*gaap',
            r'(?i)gaap\s+(?:general\s+and\s+administrative|g&a|sg&a|selling[\s,]+general[\s,]+and[\s,]+administrative)',
            r'(?i)(?:general\s+and\s+administrative|g&a|sg&a|selling[\s,]+general[\s,]+and[\s,]+administrative).*gaap',
            r'(?i)gaap\s+(?:cost\s+of\s+(?:revenue|sales|goods\s+sold)|cogs)',
            r'(?i)(?:cost\s+of\s+(?:revenue|sales|goods\s+sold)|cogs).*gaap',
            r'(?i)gaap\s+(?:selling\s+and\s+marketing|s&m)',
            r'(?i)(?:selling\s+and\s+marketing|s&m).*gaap'
        ]
        
        # Patterns for Non-GAAP line items
        non_gaap_patterns = [
            r'(?i)non[\s-]gaap\s+(?:net\s+)?(?:income|loss|earnings|profit)(?:\s+\(loss\))?',
            r'(?i)adjusted\s+(?:net\s+)?(?:income|loss|earnings|profit)(?:\s+\(loss\))?',
            r'(?i)non[\s-]gaap\s+(?:income|loss)(?:\s+from\s+operations)?(?:\s+\(loss\))?',
            r'(?i)adjusted\s+(?:income|loss)(?:\s+from\s+operations)?(?:\s+\(loss\))?',
            r'(?i)non[\s-]gaap\s+(?:gross\s+)?profit',
            r'(?i)adjusted\s+(?:gross\s+)?profit',
            r'(?i)non[\s-]gaap\s+operating\s+(?:income|loss|profit|margin)(?:\s+\(loss\))?',
            r'(?i)adjusted\s+operating\s+(?:income|loss|profit|margin)(?:\s+\(loss\))?',
            r'(?i)non[\s-]gaap\s+(?:loss|income)\s+from\s+operations',
            r'(?i)adjusted\s+(?:loss|income)\s+from\s+operations',
            r'(?i)non[\s-]gaap\s+revenue',
            r'(?i)adjusted\s+revenue',
            r'(?i)non[\s-]gaap\s+(?:loss|earnings|income)\s+per\s+share(?:\s*,\s*(?:basic|diluted))?',
            r'(?i)adjusted\s+(?:loss|earnings|income)\s+per\s+share(?:\s*,\s*(?:basic|diluted))?',
            r'(?i)non[\s-]gaap\s+(?:eps|earnings\s+per\s+share|net\s+(?:income|loss)\s+per\s+share)(?:\s*,\s*(?:basic|diluted))?',
            r'(?i)adjusted\s+(?:eps|earnings\s+per\s+share|net\s+(?:income|loss)\s+per\s+share)(?:\s*,\s*(?:basic|diluted))?',
            r'(?i)non[\s-]gaap\s+(?:basic|diluted)\s+(?:eps|earnings\s+per\s+share|(?:income|loss)\s+per\s+share)',
            r'(?i)adjusted\s+(?:basic|diluted)\s+(?:eps|earnings\s+per\s+share|(?:income|loss)\s+per\s+share)',
            r'(?i)non[\s-]gaap\s+(?:total\s+)?operating\s+(?:expenses?|costs?)',
            r'(?i)adjusted\s+(?:total\s+)?operating\s+(?:expenses?|costs?)',
            r'(?i)non[\s-]gaap\s+(?:research\s+and\s+development|r&d|r\s*&\s*d|research\s+&\s+development|research\s+development|product\s+development)',
            r'(?i)adjusted\s+(?:research\s+and\s+development|r&d|r\s*&\s*d|research\s+&\s+development|research\s+development|product\s+development)',
            r'(?i)non[\s-]gaap\s+(?:sales\s+and\s+marketing|marketing)',
            r'(?i)adjusted\s+(?:sales\s+and\s+marketing|marketing)',
            r'(?i)non[\s-]gaap\s+(?:general\s+and\s+administrative|g&a|sg&a|selling[\s,]+general[\s,]+and[\s,]+administrative)',
            r'(?i)adjusted\s+(?:general\s+and\s+administrative|g&a|sg&a|selling[\s,]+general[\s,]+and[\s,]+administrative)',
            r'(?i)non[\s-]gaap\s+(?:cost\s+of\s+(?:revenue|sales|goods\s+sold)|cogs)',
            r'(?i)adjusted\s+(?:cost\s+of\s+(?:revenue|sales|goods\s+sold)|cogs)',
            r'(?i)non[\s-]gaap\s+(?:selling\s+and\s+marketing|s&m)',
            r'(?i)adjusted\s+(?:selling\s+and\s+marketing|s&m)'
        ]
        
        # Patterns for adjustment items to exclude
        adjustment_patterns = [
            r'(?i)stock[\s-]based\s+compensation',
            r'(?i)share[\s-]based\s+compensation',
            r'(?i)amortization\s+of\s+intangible',
            r'(?i)acquisition[\s-]related',
            r'(?i)restructuring\s+(?:costs|charges)',
            r'(?i)impairment\s+(?:charges|losses)',
            r'(?i)litigation\s+(?:costs|settlements)',
            r'(?i)tax\s+(?:effects|adjustments)',
            r'(?i)one[\s-]time\s+(?:charges|items)',
            r'(?i)non[\s-]recurring\s+(?:charges|items)',
            r'(?i)depreciation\s+and\s+amortization',
            r'(?i)foreign\s+exchange',
            r'(?i)interest\s+(?:income|expense)',
            r'(?i)other\s+(?:income|expense)',
            r'(?i)provision\s+for\s+income\s+tax',
            r'(?i)income\s+tax\s+(?:benefit|expense)',
            r'(?i)change\s+in\s+fair\s+value',
            r'(?i)warrant\s+(?:liability|expense)',
            r'(?i)convertible\s+note',
            r'(?i)debt\s+(?:issuance|modification)',
            r'(?i)transaction\s+(?:costs|expenses)',
            r'(?i)professional\s+(?:fees|services)',
            r'(?i)severance\s+(?:costs|expenses)',
            r'(?i)lease\s+(?:impairment|modification)',
            r'(?i)asset\s+(?:impairment|write-off)',
            r'(?i)goodwill\s+impairment'
        ]
        
        for line_item, values in data.items():
            line_item_lower = line_item.lower()
            
            # Check if it's a GAAP line item
            is_gaap = any(re.search(pattern, line_item) for pattern in gaap_patterns)
            
            # Check if it's a Non-GAAP line item
            is_non_gaap = any(re.search(pattern, line_item) for pattern in non_gaap_patterns)
            
            # Check if it's an adjustment item (to exclude)
            is_adjustment = any(re.search(pattern, line_item) for pattern in adjustment_patterns)
            
            # Check if line item has any meaningful values (not empty, null, or zero)
            has_values = False
            if isinstance(values, list):
                has_values = any(v is not None and str(v).strip() != '' and str(v).strip() != '0' for v in values)
            elif values is not None and str(values).strip() != '' and str(values).strip() != '0':
                has_values = True
            
            # Include only GAAP and Non-GAAP items with values, exclude adjustments and empty items
            if (is_gaap or is_non_gaap) and not is_adjustment and has_values:
                filtered_data[line_item] = values
        
        return filtered_data
    
    def create_statement_dataframe_without_cleaning(self, statement_data: Dict, statement_name: str) -> pd.DataFrame:
        """Convert statement data to pandas DataFrame with SEC filing links (without cleaning)"""
        if not statement_data.get('periods') or not statement_data.get('data'):
            return pd.DataFrame()
        
        periods = statement_data['periods']
        data = statement_data['data']
        filing_links = statement_data.get('filing_links', {})
        
        # Create DataFrame without cleaning (data is already cleaned)
        df_data = {'Line Item': list(data.keys())}
        
        for i, period in enumerate(periods):
            period_values = []
            for line_item in data.keys():
                values = data[line_item]
                # Handle case where values might be a dict or other non-list type
                if isinstance(values, dict):
                    # Skip this line item as it's malformed
                    period_values.append('')
                elif isinstance(values, list):
                    if i < len(values) and values[i] is not None:
                        period_values.append(values[i])
                    else:
                        period_values.append('')
                else:
                    # Single value, use it for first period only
                    if i == 0:
                        period_values.append(values)
                    else:
                        period_values.append('')
            df_data[period] = period_values
        
        df = pd.DataFrame(df_data)
        
        # Add SEC filing links row if available
        if filing_links:
            try:
                with st.expander("🔍 Debug: Excel Export Filing Links", expanded=False):
                    st.write(f"Periods: {periods}")
                    st.write(f"Filing Links Dict: {filing_links}")
            except:
                pass
                
            link_row = {'Line Item': 'SEC Filing Link'}
            for period in periods:
                url = filing_links.get(period, '')
                try:
                    with st.expander("🔍 Debug: Period URL Mapping", expanded=False):
                        st.write(f"Period '{period}' -> URL: '{url}'")
                except:
                    pass
                # Store URL for hyperlink formatting later - ensure it's a string
                link_row[period] = str(url) if url else ''
            
            try:
                with st.expander("🔍 Debug: Final Link Row", expanded=False):
                    st.write(f"Final Link Row: {link_row}")
            except:
                pass
                
            # Add the link row as the first row after headers
            link_df = pd.DataFrame([link_row])
            df = pd.concat([link_df, df], ignore_index=True)
            
            try:
                with st.expander("✅ Debug: DataFrame with Filing Links", expanded=False):
                    st.write(f"Added SEC Filing Link row to DataFrame")
                    st.write(f"DataFrame shape after adding links: {df.shape}")
                    st.write(f"First few rows:")
                    st.dataframe(df.head(3))
            except:
                pass
        
        return df
    
    def format_worksheet(self, worksheet, df: pd.DataFrame = None):
        """Apply professional formatting to worksheet"""
        if df is not None and df.empty:
            return
        
        if df is not None:
            # Set column widths - make Line Item column much wider
            worksheet.column_dimensions['A'].width = 50
            for col in range(2, len(df.columns) + 1):
                worksheet.column_dimensions[chr(64 + col)].width = 18
            
            # Format header row
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_alignment
                cell.border = self.border
            
            # Set row heights for better readability
            for row in range(1, len(df) + 2):
                worksheet.row_dimensions[row].height = 25
            
            # Format data rows
            for row in range(2, len(df) + 2):
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = self.border
                    
                    # Check if this is the SEC Filing Link row (first data row) and has a URL
                    if row == 2 and col > 1 and cell.value and isinstance(cell.value, str) and cell.value.strip() and cell.value.startswith('http'):
                        # Format as hyperlink with the actual URL
                        cell.hyperlink = cell.value.strip()
                        cell.value = "View 8-K Filing"
                        cell.style = "Hyperlink"
                        cell.alignment = Alignment(horizontal='center')
                    elif col > 1:  # Skip line item column
                        cell.alignment = Alignment(horizontal='right')
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = self.currency_format
                    else:
                        cell.alignment = Alignment(horizontal='left')
            
            # Add alternating row colors
            light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            for row in range(2, len(df) + 2):
                if row % 2 == 0:
                    for col in range(1, len(df.columns) + 1):
                        worksheet.cell(row=row, column=col).fill = light_fill
        else:
            # Basic formatting for raw data sheets without DataFrame
            # Set column widths
            for col_idx in range(1, 50):  # Format first 50 columns
                col_letter = get_column_letter(col_idx)
                if col_idx == 1:
                    worksheet.column_dimensions[col_letter].width = 30
                else:
                    worksheet.column_dimensions[col_letter].width = 15
    
    def create_summary_sheet(self, financial_data: Dict) -> pd.DataFrame:
        """Create a summary sheet with key metrics"""
        summary_data = []
        
        # Handle the case where financial_data contains lists
        converted_data = self.convert_app_data_format(financial_data)
        
        # Get periods from income statement
        periods = converted_data.get('income_statement', {}).get('periods', [])
        if not periods:
            return pd.DataFrame()
        
        # Extract key metrics
        income_data = converted_data.get('income_statement', {}).get('data', {})
        balance_data = converted_data.get('balance_sheet', {}).get('data', {})
        cash_flow_data = converted_data.get('cash_flow', {}).get('data', {})
        
        # Revenue metrics
        if 'Total Revenue' in income_data:
            summary_data.append(['Total Revenue'] + income_data['Total Revenue'][:len(periods)])
        
        # Profitability metrics
        if 'Net Income' in income_data:
            summary_data.append(['Net Income'] + income_data['Net Income'][:len(periods)])
        
        if 'Operating Income' in income_data:
            summary_data.append(['Operating Income'] + income_data['Operating Income'][:len(periods)])
        
        # Balance sheet metrics
        if 'Total Assets' in balance_data:
            summary_data.append(['Total Assets'] + balance_data['Total Assets'][:len(periods)])
        
        if 'Total Shareholders Equity' in balance_data:
            summary_data.append(['Total Shareholders Equity'] + balance_data['Total Shareholders Equity'][:len(periods)])
        
        # Cash flow metrics
        if 'Net Cash from Operating Activities' in cash_flow_data:
            summary_data.append(['Operating Cash Flow'] + cash_flow_data['Net Cash from Operating Activities'][:len(periods)])
        
        # EPS
        if 'Diluted EPS' in income_data:
            summary_data.append(['Diluted EPS'] + income_data['Diluted EPS'][:len(periods)])
        
        # Create DataFrame
        if summary_data:
            columns = ['Metric'] + periods
            df = pd.DataFrame(summary_data, columns=columns)
            return df
        
        return pd.DataFrame()
    
    def create_sbc_breakdown_sheet(self, financial_data: Dict) -> pd.DataFrame:
        """Create a stock-based compensation breakdown sheet"""
        sbc_data = []
        
        # Convert app data format
        converted_data = self.convert_app_data_format(financial_data)
        
        # Get periods from income statement
        periods = converted_data.get('income_statement', {}).get('periods', [])
        if not periods:
            return pd.DataFrame()
        
        # Extract stock-based compensation line items from income statement
        income_data = converted_data.get('income_statement', {}).get('data', {})
        
        # Look for stock-based compensation related line items
        sbc_keywords = [
            'stock-based compensation', 'stock based compensation', 'share-based compensation',
            'share based compensation', 'equity compensation', 'stock compensation',
            'cost of revenue: subscription', 'cost of revenue: services',
            'sales and marketing', 'research and development', 'general and administrative'
        ]
        
        for line_item, values in income_data.items():
            line_item_lower = line_item.lower()
            # Check if this line item contains stock-based compensation breakdown
            if any(keyword in line_item_lower for keyword in sbc_keywords):
                # Only include if it looks like a breakdown (has specific categories)
                if ('cost of revenue' in line_item_lower or 
                    'sales and marketing' in line_item_lower or
                    'research and development' in line_item_lower or
                    'general and administrative' in line_item_lower):
                    sbc_data.append([line_item] + values[:len(periods)])
        
        # Create DataFrame if we have SBC data
        if sbc_data:
            columns = ['Stock-Based Compensation Component'] + periods
            df = pd.DataFrame(sbc_data, columns=columns)
            return df
        
        return pd.DataFrame()
    
    def convert_app_data_format(self, app_data: Dict) -> Dict:
        """Convert app data format to expected structure"""
        converted = {}
        
        # Debug: Show input data structure
        try:
            import streamlit as st
            with st.expander("🔍 Debug: convert_app_data_format Input", expanded=False):
                st.write(f"Input app_data type: {type(app_data)}")
                st.write(f"Input app_data keys: {list(app_data.keys()) if app_data else 'None'}")
                if app_data:
                    for ticker, filings in app_data.items():
                        st.write(f"  Ticker '{ticker}': {len(filings) if filings else 0} filings")
                        if filings:
                            st.write(f"    First filing keys: {list(filings[0].keys())}")
        except:
            pass
        
        for ticker, filings_list in app_data.items():
            if filings_list and len(filings_list) > 0:
                # Only process statement types that actually exist in the data
                available_statement_types = set()
                for filing in filings_list:
                    available_statement_types.update(filing.keys())
                
                # Remove non-statement keys
                available_statement_types.discard('filing_date')
                available_statement_types.discard('financial_tables')
                
                for statement_type in available_statement_types:
                    all_periods = []
                    all_data = {}
                    
                    # First pass: collect all periods from all filings
                    for filing in filings_list:
                        if statement_type in filing and isinstance(filing[statement_type], dict):
                            statement_data = filing[statement_type]
                            if 'periods' in statement_data:
                                for period in statement_data['periods']:
                                    if period not in all_periods:
                                        all_periods.append(period)
                    
                    # Second pass: map data to correct period positions
                    for filing in filings_list:
                        if statement_type in filing and isinstance(filing[statement_type], dict):
                            statement_data = filing[statement_type]
                            filing_periods = statement_data.get('periods', [])
                            
                            if 'data' in statement_data:
                                for line_item, values in statement_data['data'].items():
                                    if line_item not in all_data:
                                        # Initialize with None for all periods
                                        all_data[line_item] = [None] * len(all_periods)
                                    
                                    # Map each value to its correct period position
                                    if isinstance(values, list):
                                        for i, value in enumerate(values):
                                            if i < len(filing_periods):
                                                period = filing_periods[i]
                                                if period in all_periods:
                                                    period_index = all_periods.index(period)
                                                    all_data[line_item][period_index] = value
                    
                    if all_periods and all_data:
                        # Collect filing_links from the original data
                        all_filing_links = {}
                        for filing in filings_list:
                            if statement_type in filing and isinstance(filing[statement_type], dict):
                                filing_links = filing[statement_type].get('filing_links', {})
                                all_filing_links.update(filing_links)
                        
                        converted[statement_type] = {
                            'periods': all_periods,
                            'data': all_data,
                            'filing_links': all_filing_links
                        }
        
        # Debug: Show what we're returning
        try:
            import streamlit as st
            with st.expander("🔍 Debug: convert_app_data_format Output", expanded=False):
                st.write(f"Converted keys: {list(converted.keys())}")
                st.write(f"Converted type: {type(converted)}")
                for statement_type, data in converted.items():
                    filing_links = data.get('filing_links', {})
                    st.write(f"  {statement_type}: {len(filing_links)} filing_links")
                    if filing_links:
                        st.write(f"    Sample links: {list(filing_links.items())[:2]}")
                if not converted:
                    st.error("❌ convert_app_data_format is returning empty dict!")
        except:
            pass
        
        return converted
    
    def convert_app_data_format_raw(self, app_data: Dict) -> Dict:
        """This method is no longer used for raw data - raw data is processed directly from filings"""
        # This method is kept for compatibility but not used for raw data processing
        # Raw data is now processed directly in the Excel export without any conversion
        return {}
    
    def _clean_and_merge_line_items(self, data: Dict, periods: List) -> Dict:
        """Clean up line item names and merge semantically identical ones"""
        import openai
        import json
        import os
        
        try:
            client = openai.OpenAI(api_key=os.getenv('OPENAI_API_KEY'))
            
            line_items = list(data.keys())
            
            merge_prompt = f"""
You are a financial analyst. I have financial statement line items that need to be cleaned and merged. Be aggressive in merging semantically identical items, but ALWAYS keep GAAP and Non-GAAP versions separate.

CRITICAL REQUIREMENT: You must preserve exact column positions. Each line item has values in specific columns that correspond to specific time periods (e.g., Q1 2024, Q2 2024, etc.). When merging items, the values MUST stay in their original column positions to maintain period alignment.

DO NOT rearrange, reorder, or move values between columns. The column order represents chronological periods and must be preserved exactly.

Line items to process:
{json.dumps(list(data.keys()), indent=2)}

Sample data structure (values are in column order - DO NOT CHANGE ORDER):
{json.dumps({k: v for k, v in list(data.items())[:3]}, indent=2)}

Instructions:
1. Group semantically identical line items together
2. Keep GAAP and Non-GAAP versions separate (create separate groups)
3. Look for semantic equivalence, synonyms, and alternative phrasings
4. Clean up prefixes and standardize naming
5. Merge operating expenses with different descriptions but same function
6. Merge per-share metrics regardless of calculation method differences (Please distinguish between Diluted per share metrics and Fully Diluted per share metrics)
7. GAAP identification: Items containing "on a GAAP basis", "GAAP", or similar should be labeled as "- GAAP"
8. Common synonyms to merge:
   - Net Income = Net Loss (opposite signs)
   - Operating Income = Loss from Operations = Operating Loss (all same concept, different signs)
   - Operations = Operating Income/Loss (standardize to "Operating Income")
   - Diluted = Fully Diluted
   - Revenue = Sales = Net Sales = Total Revenue
   - Cost of Revenue = Cost of Sales = Cost of Goods Sold
   - Research and Development = R&D
   - General and Administrative = G&A
   - Sales and Marketing = Marketing
   - Interest Income = Interest Revenue
   - Interest Expense = Interest Cost
   - Income Tax = Tax Expense = Provision for Income Tax
   - Earnings Per Share = EPS
   - Weighted Average Shares = Shares Outstanding (when in same context)
9. PRESERVE EXACT COLUMN POSITIONS - do not move values between columns

CRITICAL: For each cleaned standard name, you MUST include the EXACT original line item names from the input list in the array. Do not create new names - use the exact strings from the "Line items to process" list above.

Example format:
{{
    "Revenue - GAAP": ["Total Revenue - Revenue on a GAAP basis", "Revenue - GAAP revenue"],
    "Revenue - Non-GAAP": ["Total Revenue - Non-GAAP revenue", "Revenue - Adjusted revenue"],
    "Net Income - GAAP": ["Net Loss - Net income (loss) on a GAAP basis", "Net Income - GAAP"],
    "Operating Income - GAAP": ["Loss from Operations - Loss from operations on a GAAP basis", "Operating Income - GAAP"],
    "Diluted Shares Outstanding": ["Diluted Shares Outstanding", "Weighted Average Diluted Shares"],
    "Basic Shares Outstanding": ["Basic Shares Outstanding", "Weighted Average Basic Shares"]
}}

Be aggressive in merging - if two items represent the same financial concept, merge them.
REMEMBER: Column positions represent time periods and must be preserved exactly.

IMPORTANT: Return ONLY a valid JSON object. Do not include any explanatory text, comments, or markdown formatting.
"""

            response = client.chat.completions.create(
                model="gpt-4",
                messages=[
                    {"role": "system", "content": "You are a financial data analyst. You must respond with valid JSON only. Do not include any explanatory text, markdown formatting, or code blocks. Return only the JSON object."},
                    {"role": "user", "content": merge_prompt}
                ],
                temperature=0
            )
            
            # Track token usage
            if hasattr(response, 'usage'):
                self.last_token_usage = {
                    'prompt_tokens': response.usage.prompt_tokens,
                    'completion_tokens': response.usage.completion_tokens,
                    'total_tokens': response.usage.total_tokens,
                    'model': 'gpt-4'
                }
            
            result = response.choices[0].message.content.strip()
            
            # Validate we have content
            if not result or result.isspace():
                raise ValueError("OpenAI returned empty response")
            
            merge_mapping = json.loads(result)
            
            # Apply the mapping to actually merge the data
            merged_data = {}
            
            for standard_name, original_names in merge_mapping.items():
                # Find the first original item that exists in data
                for original_name in original_names:
                    if original_name in data:
                        merged_data[standard_name] = data[original_name]
                        break
            
            # Add any unmapped original items to preserve all data
            for original_key in data.keys():
                # Check if this original key was used in any mapping
                key_was_mapped = False
                for standard_name, original_names in merge_mapping.items():
                    if original_key in original_names:
                        key_was_mapped = True
                        break
                
                if not key_was_mapped:
                    # Keep the original item to prevent data loss
                    merged_data[original_key] = data[original_key]
            
            # Only show critical errors
            if not merged_data:
                try:
                    st.error("❌ **No items matched in OpenAI response**")
                except:
                    pass
            
            return merged_data
            
        except Exception as e:
            try:
                st.error(f"❌ **OpenAI API Error:** {str(e)}")
                st.warning("🔄 **Fallback:** Using original unmerged data")
            except:
                pass
            print(f"Error merging line items: {e}")
            # Fallback to original data
            return data
    
    def export_to_excel(self, financial_data: Dict, ticker: str, include_raw_data: bool = False, raw_financial_data: Dict = None) -> bytes:
        """Export financial data to Excel file and return as bytes"""
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Convert app data structure to exporter expected format
        converted_data = self.convert_app_data_format(financial_data)
        
        # Store raw data before cleaning if requested
        raw_data = None
        if include_raw_data and raw_financial_data:
            import copy
            # Convert raw financial data to the expected format without any processing
            raw_data = self.convert_app_data_format_raw(raw_financial_data)
        
        # Check if conversion was successful
        if not converted_data:
            try:
                import streamlit as st
                with st.expander("❌ Debug: No Data to Export", expanded=True):
                    st.error("convert_app_data_format returned None")
                    st.write(f"Input financial_data keys: {list(financial_data.keys()) if financial_data else 'None'}")
            except:
                pass
            return b''  # Return empty bytes if no data
        
        # Apply cleaning and merging to all statement types after conversion
        try:
            import streamlit as st
            with st.expander("🤖 AI Line Item Cleaning & Merging", expanded=False):
                cleaning_container = st.container()
        except:
            pass
            
        for statement_type in ['income_statement', 'balance_sheet', 'cash_flow', 'sbc_breakdown', 'gaap_reconciliation']:
            if statement_type in converted_data:
                statement_data = converted_data[statement_type]
                if 'data' in statement_data and 'periods' in statement_data:
                    try:
                        with cleaning_container:
                            st.write(f"**📊 Processing {statement_type.replace('_', ' ').title()}**")
                            
                            # Show original data
                            original_items = list(statement_data['data'].keys())
                            st.write(f"📥 **Original Items ({len(original_items)}):**")
                            
                            # Display first few items in a nice format
                            for i, item in enumerate(original_items[:8]):
                                st.write(f"   {i+1}. {item}")
                            if len(original_items) > 8:
                                st.write(f"   ... and {len(original_items) - 8} more items")
                            
                            st.write("🔄 **Calling OpenAI to clean and merge...**")
                    except:
                        pass
                    
                    # Clean and merge line items using ChatGPT
                    cleaned_data = self._clean_and_merge_line_items(
                        statement_data['data'], 
                        statement_data['periods']
                    )
                    
                    # CRITICAL: Update the converted_data with cleaned results
                    converted_data[statement_type]['data'] = cleaned_data
                    
                    try:
                        with cleaning_container:
                            # Show cleaned data
                            cleaned_items = list(cleaned_data.keys())
                            st.write(f"✅ **Cleaned Items ({len(cleaned_items)}):**")
                            
                            # Display cleaned items in a nice format
                            for i, item in enumerate(cleaned_items[:8]):
                                st.write(f"   {i+1}. {item}")
                            if len(cleaned_items) > 8:
                                st.write(f"   ... and {len(cleaned_items) - 8} more items")
                            
                            # Show summary
                            reduction = len(original_items) - len(cleaned_items)
                            if reduction > 0:
                                st.success(f"🎯 **Merged {reduction} duplicate items** ({len(original_items)} → {len(cleaned_items)})")
                            else:
                                st.info(f"ℹ️ **No duplicates found** ({len(original_items)} items)")
                            
                            st.write("---")
                    except:
                        pass
        
        # Create Income Statement sheet
        income_df = self.create_statement_dataframe_without_cleaning(
            converted_data.get('income_statement', {}), 
            'Income Statement'
        )
        if not income_df.empty:
            ws_income = wb.create_sheet("Income Statement")
            for row in dataframe_to_rows(income_df, index=False, header=True):
                ws_income.append(row)
            self.format_worksheet(ws_income, income_df)
        
        # Create Balance Sheet sheet
        balance_df = self.create_statement_dataframe_without_cleaning(
            converted_data.get('balance_sheet', {}), 
            'Balance Sheet'
        )
        if not balance_df.empty:
            ws_balance = wb.create_sheet("Balance Sheet")
            for row in dataframe_to_rows(balance_df, index=False, header=True):
                ws_balance.append(row)
            self.format_worksheet(ws_balance, balance_df)
        
        # Create Cash Flow Statement sheet
        cash_flow_df = self.create_statement_dataframe_without_cleaning(
            converted_data.get('cash_flow', {}), 
            'Cash Flow Statement'
        )
        if not cash_flow_df.empty:
            ws_cash_flow = wb.create_sheet("Cash Flow Statement")
            for row in dataframe_to_rows(cash_flow_df, index=False, header=True):
                ws_cash_flow.append(row)
            self.format_worksheet(ws_cash_flow, cash_flow_df)
        
        # Create Stock-Based Compensation sheet
        sbc_df = self.create_statement_dataframe_without_cleaning(
            converted_data.get('sbc_breakdown', {}), 
            'Stock-Based Compensation'
        )
        if not sbc_df.empty:
            ws_sbc = wb.create_sheet("Stock-Based Compensation")
            for row in dataframe_to_rows(sbc_df, index=False, header=True):
                ws_sbc.append(row)
            self.format_worksheet(ws_sbc, sbc_df)
        
        # Create GAAP Reconciliation sheet (uses processed data with duplicate removal)
        gaap_df = self.create_statement_dataframe(
            converted_data.get('gaap_reconciliation', {}), 
            'GAAP Reconciliation'
        )
        if not gaap_df.empty:
            ws_gaap = wb.create_sheet("GAAP Reconciliation")
            for row in dataframe_to_rows(gaap_df, index=False, header=True):
                ws_gaap.append(row)
            self.format_worksheet(ws_gaap, gaap_df)
        
        # Create Raw Data sheet if requested (shows filings side-by-side horizontally)
        if include_raw_data and raw_financial_data:
            try:
                import streamlit as st
            except:
                pass
            
            ws_raw = wb.create_sheet("Raw Data")
            
            # Collect all filing data organized by statement type
            filing_data_by_statement = {}
            
            # Process each ticker's filings
            for ticker, filings_list in raw_financial_data.items():
                if filings_list:
                    for filing_idx, filing in enumerate(filings_list):
                        filing_date = filing.get('filing_date', f'Filing_{filing_idx+1}')
                        
                        # Process each statement type in this filing
                        for statement_type in ['income_statement', 'balance_sheet', 'cash_flow', 'gaap_reconciliation', 'sbc_breakdown']:
                            if statement_type in filing and isinstance(filing[statement_type], dict):
                                statement_data = filing[statement_type]
                                periods = statement_data.get('periods', [])
                                data = statement_data.get('data', {})
                                filing_links = statement_data.get('filing_links', {})
                                
                                if statement_type not in filing_data_by_statement:
                                    filing_data_by_statement[statement_type] = []
                                
                                filing_data_by_statement[statement_type].append({
                                    'filing_date': filing_date,
                                    'periods': periods,
                                    'data': data,
                                    'filing_links': filing_links
                                })
            
            # Create horizontal layout for each statement type
            current_row = 1
            
            for statement_type, filings in filing_data_by_statement.items():
                if not filings:
                    continue
                    
                # Add statement type header with enhanced formatting
                header_cell = ws_raw.cell(row=current_row, column=1, value=f"{statement_type.replace('_', ' ').title()}")
                header_cell.font = Font(bold=True, size=16, color="FFFFFF")
                header_cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
                header_cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Merge header across multiple columns for better visibility
                ws_raw.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
                current_row += 2
                
                # Create column headers for each filing
                col_offset = 1
                for filing in filings:
                    filing_date = filing['filing_date']
                    periods = filing['periods']
                    
                    # Add filing date header spanning across all columns for this filing
                    merge_end_col = col_offset + len(periods)
                    ws_raw.merge_cells(start_row=current_row, start_column=col_offset, 
                                     end_row=current_row, end_column=merge_end_col)
                    filing_header = ws_raw.cell(row=current_row, column=col_offset, value=f"Filing: {filing_date}")
                    filing_header.font = Font(bold=True, color="FFFFFF")
                    filing_header.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    filing_header.alignment = Alignment(horizontal='center', vertical='center')
                    filing_header.border = self.border
                    
                    # Add "Line Item" header with enhanced styling
                    line_item_header = ws_raw.cell(row=current_row + 1, column=col_offset, value="Line Item")
                    line_item_header.font = self.header_font
                    line_item_header.fill = self.header_fill
                    line_item_header.alignment = self.center_alignment
                    line_item_header.border = self.border
                    
                    # Add ALL period headers for this filing with enhanced styling
                    for i, period in enumerate(periods):
                        period_header = ws_raw.cell(row=current_row + 1, column=col_offset + 1 + i, value=period)
                        period_header.font = self.header_font
                        period_header.fill = self.header_fill
                        period_header.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                        period_header.border = self.border
                    
                    # Set row height for period header row to accommodate wrapped text
                    ws_raw.row_dimensions[current_row + 1].height = 40
                    
                    # Move to next filing position (line item column + ALL period columns + empty column)
                    col_offset += len(periods) + 2  # +1 for line item column, +1 for empty separator
                
                
                # Add SEC Filing Link row for each filing
                link_row = current_row + 2
                col_offset = 1
                
                for filing in filings:
                    filing_links = filing.get('filing_links', {})
                    periods = filing['periods']
                    
                    # Add "SEC Filing Link" label
                    ws_raw.cell(row=link_row, column=col_offset, value="SEC Filing Link")
                    ws_raw.cell(row=link_row, column=col_offset).font = Font(bold=True, color="0000FF")
                    
                    # Add filing URLs for each period
                    for i, period in enumerate(periods):
                        url = filing_links.get(period, '')
                        if url and isinstance(url, str) and url.strip() and url.startswith('http'):
                            cell = ws_raw.cell(row=link_row, column=col_offset + 1 + i, value="View 8-K Filing")
                            cell.hyperlink = url.strip()
                            cell.style = "Hyperlink"
                            cell.alignment = Alignment(horizontal='center')
                    
                    # Move to next filing position
                    col_offset += len(periods) + 2
                
                # Add data rows - each filing gets its own line items independently
                data_start_row = current_row + 3  # Account for link row
                
                # Find the maximum number of line items across all filings to determine row count
                max_line_items = max(len(filing['data']) for filing in filings) if filings else 0
                
                # Create rows for each filing independently
                for row_idx in range(max_line_items):
                    col_offset = 1
                    
                    for filing in filings:
                        filing_data = filing['data']
                        periods = filing['periods']
                        
                        # Get line items in original order for this specific filing
                        filing_line_items = list(filing_data.keys())
                        
                        # Add line item and values if this filing has a line item at this row index
                        if row_idx < len(filing_line_items):
                            line_item = filing_line_items[row_idx]
                            
                            # Add line item name with formatting
                            line_item_cell = ws_raw.cell(row=data_start_row + row_idx, column=col_offset, value=line_item)
                            line_item_cell.alignment = Alignment(horizontal='left', vertical='center')
                            line_item_cell.border = self.border
                            
                            # Add alternating row colors
                            if row_idx % 2 == 0:
                                line_item_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                            
                            # Add values for ALL periods in this line item
                            values = filing_data[line_item]
                            if isinstance(values, list):
                                for i, value in enumerate(values):
                                    if i < len(periods) and value is not None:
                                        value_cell = ws_raw.cell(row=data_start_row + row_idx, column=col_offset + 1 + i, value=str(value))
                                        value_cell.alignment = Alignment(horizontal='right', vertical='center')
                                        value_cell.border = self.border
                                        
                                        # Apply alternating row colors
                                        if row_idx % 2 == 0:
                                            value_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                                        
                                        # Format numeric values
                                        try:
                                            float_val = float(str(value).replace(',', '').replace('$', '').replace('(', '-').replace(')', ''))
                                            value_cell.number_format = self.currency_format
                                        except:
                                            pass
                        
                        # Move to next filing position (accounting for ALL periods)
                        col_offset += len(periods) + 2
                
                # Move to next statement type
                current_row = data_start_row + max_line_items + 3
            
            # Set column widths for better readability
            ws_raw.column_dimensions['A'].width = 35  # Line item column wider
            for col_idx in range(2, 50):  # Set width for data columns
                col_letter = get_column_letter(col_idx)
                ws_raw.column_dimensions[col_letter].width = 18
            
            # Apply row height for better spacing
            for row_idx in range(1, ws_raw.max_row + 1):
                ws_raw.row_dimensions[row_idx].height = 20
        
        # If no sheets were created, create a placeholder
        if len(wb.worksheets) == 0:
            ws = wb.create_sheet("No Data")
            ws['A1'] = f"No financial data found for {ticker}"
            ws['A1'].font = Font(bold=True)
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()
    
    def create_excel_export(self, financial_data: Dict, ticker: str, include_raw_data: bool = False, raw_financial_data: Dict = None) -> bytes:
        """Alias for export_to_excel method to match app.py expectations"""
        return self.export_to_excel(financial_data, ticker, include_raw_data, raw_financial_data)
